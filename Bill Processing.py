"""
   Auto-generated Bill Scraper
   Version 1.0
   Coded in Python 2.7

   Modules used:
       OpenPyXl - for handling xlsx sheets
       os - to access and change working directories
       
   Features:
       Scrapes the bill book(in the form excel sheet) to 
       list out customers who have moved out/in and those
       who have changed their monthly packages.
       The above manipulated details are neatly tabulated 
       in a new workbook with necessary customer details 
       in distinct sheets.
    
   Future Scope:
       To extend the same idea with GUI to minimize the 
       efforts from user side from entering details 
       regarding the workbook such as name, directory in 
       which it is present directly into the master code.
"""
    
import openpyxl
import os

book1=raw_input("Enter old bill book name : ")
book2=raw_input("Enter new bill book name : ")
book3=raw_input("Enter ledger name : ")

#Declaring and opening workboooks required for the program
#------------------------------------------------------------------------
prev_month = openpyxl.load_workbook(book1+".xlsx")
book = openpyxl.load_workbook(book3+".xlsx")
next_month = openpyxl.load_workbook(book2+".xlsx")
#------------------------------------------------------------------------

#Creating and opening sheets needed for manipulation
#------------------------------------------------------------------------
next_month.create_sheet(index=1,title="New boxes")
next_month.create_sheet(index=2,title="Disconnected")
next_month.create_sheet(index=3,title="Plan changed")
ledger = book.get_sheet_by_name("Main")
old = prev_month.get_sheet_by_name("Sheet1")
new = next_month.get_sheet_by_name("Sheet1")
new_connection = next_month.get_sheet_by_name("New boxes")
cut = next_month.get_sheet_by_name("Disconnected")
plan = next_month.get_sheet_by_name("Plan changed")
#------------------------------------------------------------------------

#Naming the columns/attributes for respective sheets in the workbooks
#------------------------------------------------------------------------
cut['A1'],cut['B1'],cut['C1']="Disconnected STB","Amount","Street"
new_connection['A1'],new_connection['B1'],new_connection['C1']="New STB","Amount","Street"
plan['A1'],plan['B1'],plan['C1'],plan['D1'],plan['E1']="STB No"," Old Amount","New Amount","Street","Package"
#------------------------------------------------------------------------

#list for tracking disconnected boxes
list_of_disconnected_boxes=[]

#list for tracking new boxes
list_of_new_boxes=[]

row=2

# code for disconnected sheets in "bettered" book of the month
#------------------------------------------------------------------------
for i in range(3,old.max_row-1):
    # fetching details of old box
    old_box = old['G'+str(i)].value
    old_amt = old['F'+str(i)].value
    box_found = False
    
    #parsing through the new bill checking for match
    for j in range(3,new.max_row-1):
        new_box = new['G'+str(j)].value
        new_amt = new['F'+str(j)].value
        if old_box==new_box:
            box_found=True
            break
    
    if box_found == False:
        #parsing through the ledger to find street
        for j in range(2,ledger.max_row+1):
            if old_box==ledger['D'+str(j)].value:
                cut['C'+str(row)]=ledger['E'+str(j)].value
                break
        #updating the details of disconnected box in the new book
        cut['A'+str(row)]=old_box
        cut['B'+str(row)]=old_amt
        list_of_disconnected_boxes.append(old_box)
        row+=1
#------------------------------------------------------------------------
        
row=2

# code for New boxes sheet in "bettered" book of the month
#------------------------------------------------------------------------
for i in range(3,new.max_row-1):            
    # fetching details of new box
    new_box = new['G'+str(i)].value
    new_amt = new['F'+str(i)].value
    box_found = False
    #parsing through the old bill checking for match
    for j in range(3,old.max_row-1):
        old_box = old['G'+str(j)].value
        old_amt = old['F'+str(j)].value
        
        if old_box==new_box:
            box_found=True
            break
            
    if box_found == False:
        #updating the details of new box in the new book
        for j in range(2,ledger.max_row+1):
            if new_box==ledger['D'+str(j)].value:
                new_connection['C'+str(row)]=ledger['E'+str(j)].value
                break
        new_connection['A'+str(row)]=new_box
        new_connection['B'+str(row)]=new_amt
        list_of_new_boxes.append(new_box)
        row+=1
#------------------------------------------------------------------------

ct=2

# code for plan change sheets in "bettered" book of the month
#------------------------------------------------------------------------
for i in range(3,new.max_row-1): 
    # fetching details of new box                     
    new_box = new['G'+str(i)].value
    new_amt = new['F'+str(i)].value
    pack = new['H'+str(i)].value
    check = False
    
    #skipping the iteration if the box is a new box 
    if new_box in list_of_new_boxes:
        continue
        
    for j in range(3,old.max_row-1):
        old_box = old['G'+str(j)].value
        old_amt = old['F'+str(j)].value
         
        #skipping the iteration if the box is a disconnected box
        if old_box in list_of_disconnected_boxes :
            continue
        
        # if the amounts does not match adding the box to plan change sheet        
        elif old_box==new_box and old_amt!=new_amt:
            plan['A'+str(ct)]=new_box  
            plan['B'+str(ct)]=old_amt
            plan['C'+str(ct)]=new_amt
            plan['E'+str(ct)]=pack
            for z in range(2,ledger.max_row+1):
                if new_box==ledger['D'+str(z)].value:
                    plan['D'+str(ct)]=ledger['E'+str(z)].value
            ct+=1   
#------------------------------------------------------------------------


#Saving the updated workbook
#------------------------------------------------------------------------
next_month.save("bettered "+book2+".xlsx")
print "Bill generated !!!\nPlease check 'bettered "+book2+"' book in 'C:\Users\dell' for the new bill.\n"           
#------------------------------------------------------------------------



"""------------------------------------------------------------------------
                                #End of code
------------------------------------------------------------------------"""