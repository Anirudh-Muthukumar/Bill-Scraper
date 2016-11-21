import openpyxl
import os

b1=raw_input("Enter bettered book name: ")
b2=raw_input("Enter ledger name : ")

book = openpyxl.load_workbook("bettered "+b1+".xlsx")
book2 = openpyxl.load_workbook(b2+" ledger.xlsx")
sheet = book.get_sheet_by_name("New boxes")
sheet1 = book.get_sheet_by_name("Plan changed")
record = book.get_sheet_by_name("Sheet1")

list_of_change=[]

for i in range(2,sheet1.max_row+1):
    list_of_change.append(sheet1['A'+str(i)].value)

ledger = book2.get_sheet_by_name("Main")
count= ledger.max_row+1


for i in range(2,sheet.max_row+1):
    new_box= sheet['A'+str(i)].value
    
    for j in range(4,record.max_row+1):
        if new_box==record['G'+str(j)].value:
            ledger['A'+str(count)]=record['B'+str(j)].value
            ledger['B'+str(count)]=record['C'+str(j)].value
            ledger['C'+str(count)]=record['H'+str(j)].value
            ledger['D'+str(count)]=record['G'+str(j)].value
            count+=1
         
for i in range(4,record.max_row-1):
    box= record['G'+str(i)].value
    amt= record['F'+str(i)].value
    
    for j in range(1,ledger.max_row+1):
        if box==ledger['D'+str(j)].value:
            ledger['F'+str(j)]=amt

for i in range(1,ledger.max_row+1):
    box=ledger['D'+str(i)].value
    if box in list_of_change:
        for j in range(4,record.max_row-1):
            if box==record['G'+str(j)].value:
                ledger['C'+str(i)]=record['H'+str(j)].value
        
        
    
print "New ledger "+b1+" ledger is created in "+os.getcwd()   
book2.save(b1+" ledger.xlsx")

print "Success"
        
        




